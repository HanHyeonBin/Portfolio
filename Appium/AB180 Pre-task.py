import time #sleep 을 활용하기 위한 time import
import datetime #엑셀명 저장을 위한 datetime import
import random #0~100까지의 랜덤한 값을 입력하기 위해 random import
from openpyxl import Workbook #엑셀 생성을 위한 openpyxl-Workbook import
from appium import webdriver # appium webdriver를 통해 실행하기 때문에 import
from appium.webdriver.common.mobileby import MobileBy #element의 id를 기준으로 찾기 위한 mobileBy import

desired_cap = {
    "appium:deviceName" : "emulator-5554",
    "platformName" : "Android",
    "appium:appPackage" : "com.assignment.qa",
    "appium:appActivity" : "com.assignment.qa.MainActivity"
}

# 엑셀 파일 생성 및 시트 선택
wb = Workbook()
ws = wb.active


# App 실행 -> 각 입력 부분 객체 선택 -> 1~100 랜덤한 정수 값 입력 -> buyit(API호출) 버튼 선택 -> 버튼 선택에 대한 결과 값 출력을 위한 반복문
for i in range(1, 11):
    driver = webdriver.Remote("http://localhost:4723/wd/hub", desired_cap) #WebDriver 생성하여 앱 실행

    random_number = random.randint(1,100) #1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # beer_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    beer_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/beer_input")
    beer_input.click()
    time.sleep(1)
    beer_input.clear()
    time.sleep(1)
    beer_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # cookie_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    cookie_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/cookie_input")
    cookie_input.click()
    time.sleep(1)
    cookie_input.clear()
    time.sleep(1)
    cookie_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # cocoa_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    cocoa_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/cocoa_input")
    cocoa_input.click()
    time.sleep(1)
    cocoa_input.clear()
    time.sleep(1)
    cocoa_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # coke_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    coke_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/coke_input")
    coke_input.click()
    time.sleep(1)
    coke_input.clear()
    time.sleep(1)
    coke_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # orange_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    orange_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/orange_input")
    orange_input.click()
    time.sleep(1)
    orange_input.clear()
    time.sleep(1)
    orange_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # watermelon_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    watermelon_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/watermelon_input")
    watermelon_input.click()
    time.sleep(1)
    watermelon_input.clear()
    time.sleep(1)
    watermelon_input.send_keys(str(random_number))
    driver.hide_keyboard()

    random_number = random.randint(1, 100)  # 1~100까지의 정수 값 랜덤으로 random_number 변수에 저장

    # banana_input을 선택하여 랜덤한 정수 값 입력 후 키보드 닫는 코드
    banana_input = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/banana_input")
    banana_input.click()
    time.sleep(1)
    banana_input.clear()
    time.sleep(1)
    banana_input.send_keys(str(random_number))
    driver.hide_keyboard()

    # buyit(API호출) 버튼 선택
    buybtn = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/buy_button")
    buybtn.click()

    time.sleep(10)

    # buyit(API호출) 버튼 선택하여 나온 결과 값 텍스트 result 변수에 저장 후 result 텍스트 값 호출
    price_result = driver.find_element(by=MobileBy.ID, value="com.assignment.qa:id/price_result")
    result = price_result.text
    print(result)

    # 결과 Slack 전달을 위한 엑셀에 반복된 회차와 해당 회차의 api 호출 결과 텍스트 저장
    ws.cell(row=i, column=1, value=str(i))
    ws.cell(row=i, column=2, value=result)

    time.sleep(3)

# 저장할 엑셀명에 현재 시간 값을 포함하기 위한 now 변수에 저장
now = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# 저장할 엑셀명 설정과 엑셀 저장 경로
# 엑셀 저장 경로 설정 시 han11은 각 PC의 설정명으로 변경 필요
filename = "result" + now + ".xlsx"
filepath = "C:/Users/han11/Desktop/" + filename

wb.save(filepath)


driver.quit()






